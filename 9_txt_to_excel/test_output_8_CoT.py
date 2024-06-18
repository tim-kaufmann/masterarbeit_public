import openpyxl
from output_8_CoT import txt2excle

class Testtxt2excle:
    def test_txt2excle_2(self, tmp_path):
        file_path = tmp_path / "test2.txt"
        file_path.write_text("apple\tbanana\tcherry")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A1'].value == 'apple'
        assert sheet['B1'].value == 'banana'
        assert sheet['C1'].value == 'cherry'

    def test_txt2excle_3(self, tmp_path):
        file_path = tmp_path / "test3.txt"
        file_path.write_text("100\t200\t300\t400\t500")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A1'].value == '100'
        assert sheet['B1'].value == '200'
        assert sheet['C1'].value == '300'
        assert sheet['D1'].value == '400'
        assert sheet['E1'].value == '500'

    def test_txt2excle_4(self, tmp_path):
        file_path = tmp_path / "test4.txt"
        file_path.write_text("a\tb\tc\nd\te\tf")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A1'].value == 'a'
        assert sheet['B1'].value == 'b'
        assert sheet['C1'].value == 'c'
        assert sheet['A2'].value == 'd'
        assert sheet['B2'].value == 'e'
        assert sheet['C2'].value == 'f'

    def test_txt2excle_5(self, tmp_path):
        file_path = tmp_path / "test5.txt"
        file_path.write_text("1.1\t2.2\t3.3")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A1'].value == '1.1'
        assert sheet['B1'].value == '2.2'
        assert sheet['C1'].value == '3.3'

    def test_txt2excle_6(self, tmp_path):
        file_path = tmp_path / "test6.txt"
        file_path.write_text("True\tFalse")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A1'].value == 'True'
        assert sheet['B1'].value == 'False'

    def test_txt2excle_7(self, tmp_path):
        file_path = tmp_path / "test7.txt"
        file_path.write_text("")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A1'].value == None

    def test_txt2excle_8(self, tmp_path):
        file_path = tmp_path / "test8.txt"
        file_path.write_text("\t\t\t")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A1'].value == None
        assert sheet['B1'].value == None
        assert sheet['C1'].value == None

    def test_txt2excle_9(self, tmp_path):
        file_path = tmp_path / "test9.txt"
        file_path.write_text("a\tb\tc\nd\te\tf\ng\th\ti")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A3'].value == 'g'
        assert sheet['B3'].value == 'h'
        assert sheet['C3'].value == 'i'

    def test_txt2excle_10(self, tmp_path):
        file_path = tmp_path / "test10.txt"
        file_path.write_text("1\t2\t3\n4\t5\t6\n7\t8\t9")
        txt2excle(file_path)
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active
        assert sheet['A3'].value == '7'
        assert sheet['B3'].value == '8'
        assert sheet['C3'].value == '9'
